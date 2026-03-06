"""Excel and CSV spreadsheet processing (BL-267).

Uses openpyxl for .xlsx files and the csv module for .csv files.
Supports three extraction strategies:
  1. Full table (small sheets <50 rows) -> markdown table
  2. Summary (large sheets >=50 rows) -> stats + sample rows
  3. Schema-based extraction -> structured JSON rows
"""

from __future__ import annotations

import csv
import io
import logging
import math
from dataclasses import dataclass, field
from typing import Optional

logger = logging.getLogger(__name__)

# Row threshold: sheets with fewer rows get full markdown tables
SMALL_SHEET_THRESHOLD = 50

# Maximum characters for markdown output (token budget)
MAX_OUTPUT_CHARS = 8000

# Number of sample rows for large sheet summaries
SAMPLE_HEAD_ROWS = 5
SAMPLE_TAIL_ROWS = 3


@dataclass
class SheetInfo:
    """Metadata for a single worksheet."""

    name: str
    row_count: int
    col_count: int
    headers: list[str] = field(default_factory=list)


@dataclass
class SheetData:
    """Extracted data from a single worksheet."""

    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list] = field(default_factory=list)
    row_count: int = 0
    col_count: int = 0


@dataclass
class ExcelExtractionResult:
    """Complete extraction result for a spreadsheet file."""

    sheets: list[SheetInfo] = field(default_factory=list)
    markdown: str = ""
    truncated: bool = False
    errors: list[str] = field(default_factory=list)


@dataclass
class SchemaField:
    """Field definition for schema-based extraction."""

    name: str
    type: str = "string"  # string, number, boolean
    source_column: Optional[str] = None  # column header to map from


@dataclass
class SchemaExtractionResult:
    """Result of schema-based data extraction."""

    rows: list[dict] = field(default_factory=list)
    unmapped_columns: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    error: Optional[str] = None


def discover_sheets(file_path: str) -> list[SheetInfo]:
    """Discover all sheets in an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        List of SheetInfo with names and dimensions.
    """
    try:
        import openpyxl
    except ImportError:
        return []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(max_row=1, values_only=True))
            headers = [str(c) if c is not None else "" for c in rows[0]] if rows else []
            sheets.append(
                SheetInfo(
                    name=name,
                    row_count=ws.max_row or 0,
                    col_count=ws.max_column or 0,
                    headers=headers,
                )
            )
        wb.close()
        return sheets
    except Exception:
        logger.exception("Failed to discover sheets: %s", file_path)
        return []


def _read_sheet_data(file_path: str, sheet_name: Optional[str] = None) -> SheetData:
    """Read all data from a single worksheet.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Sheet to read (default: active sheet).

    Returns:
        SheetData with headers and rows.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([_clean_cell(c) for c in row])

    wb.close()

    if not all_rows:
        return SheetData(name=sheet_name or "Sheet1")

    headers = [str(c) if c else "col_{}".format(i) for i, c in enumerate(all_rows[0])]
    data_rows = all_rows[1:]

    return SheetData(
        name=sheet_name or (ws.title if ws.title else "Sheet1"),
        headers=headers,
        rows=data_rows,
        row_count=len(data_rows),
        col_count=len(headers),
    )


def read_csv_data(file_path: str) -> SheetData:
    """Read data from a CSV file.

    Args:
        file_path: Path to the CSV file.

    Returns:
        SheetData with headers and rows.
    """
    try:
        with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
            return read_csv_from_text(f.read())
    except UnicodeDecodeError:
        with open(file_path, "r", newline="", encoding="latin-1") as f:
            return read_csv_from_text(f.read())


def read_csv_from_text(text: str) -> SheetData:
    """Read data from CSV text content.

    Args:
        text: CSV file content as string.

    Returns:
        SheetData with headers and rows.
    """
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    if not all_rows:
        return SheetData(name="CSV")

    headers = [
        str(c).strip() if c else "col_{}".format(i) for i, c in enumerate(all_rows[0])
    ]
    data_rows = []
    for row in all_rows[1:]:
        data_rows.append([_clean_cell(c) for c in row])

    return SheetData(
        name="CSV",
        headers=headers,
        rows=data_rows,
        row_count=len(data_rows),
        col_count=len(headers),
    )


def read_csv_from_bytes(data: bytes) -> SheetData:
    """Read data from CSV bytes.

    Args:
        data: Raw CSV file content.

    Returns:
        SheetData with headers and rows.
    """
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    return read_csv_from_text(text)


def extract_from_file(
    file_path: str, sheet_name: Optional[str] = None
) -> ExcelExtractionResult:
    """Extract content from an Excel or CSV file.

    Applies the appropriate strategy based on sheet size:
    - Small sheets (<50 rows): full markdown table
    - Large sheets (>=50 rows): summary with stats and sample rows

    Args:
        file_path: Path to the spreadsheet file.
        sheet_name: Specific sheet to extract (Excel only).

    Returns:
        ExcelExtractionResult with markdown content.
    """
    is_csv = file_path.lower().endswith(".csv")

    if is_csv:
        return _extract_csv(file_path)

    return _extract_excel(file_path, sheet_name)


def extract_from_bytes(
    data: bytes,
    filename: str = "file.xlsx",
    sheet_name: Optional[str] = None,
) -> ExcelExtractionResult:
    """Extract content from spreadsheet bytes.

    Args:
        data: Raw file content.
        filename: Original filename (for type detection).
        sheet_name: Specific sheet to extract (Excel only).

    Returns:
        ExcelExtractionResult with markdown content.
    """
    import tempfile
    import os

    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        sheet_data = read_csv_from_bytes(data)
        return _build_result_from_sheet(sheet_data)

    # Write to temp file for openpyxl (requires seekable file)
    suffix = ".xlsx"
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name

        result = _extract_excel(tmp_path, sheet_name)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    return result


def extract_data_with_schema(
    file_path: str,
    schema_fields: list[dict],
    sheet_name: Optional[str] = None,
) -> SchemaExtractionResult:
    """Extract structured data from a spreadsheet using a schema.

    Maps spreadsheet columns to schema fields and returns typed JSON rows.

    Args:
        file_path: Path to the spreadsheet file.
        schema_fields: List of field definitions, each with:
            - name (str): output field name
            - type (str): "string", "number", or "boolean"
            - source_column (str, optional): header to map from
        sheet_name: Specific sheet (Excel only).

    Returns:
        SchemaExtractionResult with structured rows.
    """
    is_csv = file_path.lower().endswith(".csv")

    try:
        if is_csv:
            sheet_data = read_csv_data(file_path)
        else:
            try:
                import openpyxl  # noqa: F401
            except ImportError:
                return SchemaExtractionResult(
                    error="openpyxl not installed -- run: pip install openpyxl"
                )
            sheet_data = _read_sheet_data(file_path, sheet_name)
    except Exception as exc:
        return SchemaExtractionResult(error="Failed to read file: {}".format(str(exc)))

    if not sheet_data.headers:
        return SchemaExtractionResult(error="No headers found in spreadsheet")

    # Build column mapping
    fields = [SchemaField(**f) if isinstance(f, dict) else f for f in schema_fields]
    col_map = _build_column_mapping(fields, sheet_data.headers)

    # Track unmapped source columns
    mapped_headers = set(col_map.values())
    unmapped = [h for h in sheet_data.headers if h not in mapped_headers and h.strip()]

    result = SchemaExtractionResult(unmapped_columns=unmapped)

    for row_idx, row in enumerate(sheet_data.rows):
        record = {}
        for f in fields:
            source_col = col_map.get(f.name)
            if source_col is None:
                result.warnings.append(
                    "Row {}: no source column mapped for field '{}'".format(
                        row_idx + 1, f.name
                    )
                )
                record[f.name] = None
                continue

            col_idx = sheet_data.headers.index(source_col)
            if col_idx < len(row):
                raw_value = row[col_idx]
                record[f.name] = _coerce_value(
                    raw_value, f.type, f.name, row_idx + 1, result
                )
            else:
                record[f.name] = None

        result.rows.append(record)

    # Deduplicate warnings
    result.warnings = list(dict.fromkeys(result.warnings))

    return result


def sheet_to_markdown(sheet_data: SheetData) -> str:
    """Convert a SheetData to a full markdown table.

    Args:
        sheet_data: Sheet data with headers and rows.

    Returns:
        Markdown table string.
    """
    if not sheet_data.headers:
        return ""

    lines = []
    lines.append("| " + " | ".join(str(h) for h in sheet_data.headers) + " |")
    lines.append("| " + " | ".join("---" for _ in sheet_data.headers) + " |")

    for row in sheet_data.rows:
        padded = list(row) + [""] * (len(sheet_data.headers) - len(row))
        cells = [
            str(c) if c is not None else "" for c in padded[: len(sheet_data.headers)]
        ]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def sheet_to_summary(sheet_data: SheetData) -> str:
    """Convert a large SheetData to a summary with stats.

    Includes: column headers, row count, sample rows (head + tail),
    and basic stats (min/max/avg) on numeric columns.

    Args:
        sheet_data: Sheet data with headers and rows.

    Returns:
        Markdown summary string.
    """
    if not sheet_data.headers:
        return ""

    parts = []
    parts.append("**Sheet**: {}".format(sheet_data.name))
    parts.append(
        "**Dimensions**: {} rows x {} columns".format(
            sheet_data.row_count, sheet_data.col_count
        )
    )
    parts.append("**Columns**: {}".format(", ".join(sheet_data.headers)))

    # Sample rows (head)
    parts.append("")
    parts.append(
        "**Sample rows (first {}):**".format(
            min(SAMPLE_HEAD_ROWS, sheet_data.row_count)
        )
    )

    head_rows = sheet_data.rows[:SAMPLE_HEAD_ROWS]
    sample_data = SheetData(
        name=sheet_data.name,
        headers=sheet_data.headers,
        rows=head_rows,
        row_count=len(head_rows),
        col_count=sheet_data.col_count,
    )
    parts.append(sheet_to_markdown(sample_data))

    # Tail rows if enough data
    if sheet_data.row_count > SAMPLE_HEAD_ROWS + SAMPLE_TAIL_ROWS:
        tail_rows = sheet_data.rows[-SAMPLE_TAIL_ROWS:]
        parts.append("")
        parts.append("**Last {} rows:**".format(SAMPLE_TAIL_ROWS))
        tail_data = SheetData(
            name=sheet_data.name,
            headers=sheet_data.headers,
            rows=tail_rows,
            row_count=len(tail_rows),
            col_count=sheet_data.col_count,
        )
        parts.append(sheet_to_markdown(tail_data))

    # Numeric column stats
    stats = _compute_numeric_stats(sheet_data)
    if stats:
        parts.append("")
        parts.append("**Numeric column statistics:**")
        for col_name, col_stats in stats.items():
            parts.append(
                "- **{}**: min={}, max={}, avg={:.2f}, count={}".format(
                    col_name,
                    col_stats["min"],
                    col_stats["max"],
                    col_stats["avg"],
                    col_stats["count"],
                )
            )

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _clean_cell(value):
    """Clean a cell value for consistent handling."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def _extract_csv(file_path: str) -> ExcelExtractionResult:
    """Extract from a CSV file."""
    try:
        sheet_data = read_csv_data(file_path)
        return _build_result_from_sheet(sheet_data)
    except Exception as exc:
        logger.exception("CSV extraction failed: %s", file_path)
        return ExcelExtractionResult(
            errors=["CSV extraction failed: {}".format(str(exc))]
        )


def _extract_excel(
    file_path: str, sheet_name: Optional[str] = None
) -> ExcelExtractionResult:
    """Extract from an Excel file."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return ExcelExtractionResult(
            errors=["openpyxl not installed -- run: pip install openpyxl"]
        )

    try:
        sheets = discover_sheets(file_path)
        if not sheets:
            return ExcelExtractionResult(errors=["No sheets found in workbook"])

        result = ExcelExtractionResult(sheets=sheets)
        parts = []

        # Sheet index header
        if len(sheets) > 1:
            parts.append("## Workbook Sheets")
            for s in sheets:
                parts.append(
                    "- **{}**: {} rows x {} cols".format(
                        s.name, s.row_count, s.col_count
                    )
                )
            parts.append("")

        # Process requested sheet or all sheets
        sheets_to_process = (
            [s for s in sheets if s.name == sheet_name] if sheet_name else sheets
        )

        for sheet_info in sheets_to_process:
            sheet_data = _read_sheet_data(file_path, sheet_info.name)
            parts.append("## {}".format(sheet_info.name))
            parts.append("")

            if sheet_data.row_count < SMALL_SHEET_THRESHOLD:
                parts.append(sheet_to_markdown(sheet_data))
            else:
                parts.append(sheet_to_summary(sheet_data))

            parts.append("")

        markdown = "\n".join(parts)

        # Apply token budget
        if len(markdown) > MAX_OUTPUT_CHARS:
            markdown = markdown[:MAX_OUTPUT_CHARS] + "\n\n*[Content truncated]*"
            result.truncated = True

        result.markdown = markdown
        return result

    except Exception as exc:
        logger.exception("Excel extraction failed: %s", file_path)
        return ExcelExtractionResult(
            errors=["Excel extraction failed: {}".format(str(exc))]
        )


def _build_result_from_sheet(sheet_data: SheetData) -> ExcelExtractionResult:
    """Build an ExcelExtractionResult from a single SheetData."""
    sheets = [
        SheetInfo(
            name=sheet_data.name,
            row_count=sheet_data.row_count,
            col_count=sheet_data.col_count,
            headers=sheet_data.headers,
        )
    ]

    if sheet_data.row_count < SMALL_SHEET_THRESHOLD:
        markdown = sheet_to_markdown(sheet_data)
    else:
        markdown = sheet_to_summary(sheet_data)

    truncated = False
    if len(markdown) > MAX_OUTPUT_CHARS:
        markdown = markdown[:MAX_OUTPUT_CHARS] + "\n\n*[Content truncated]*"
        truncated = True

    return ExcelExtractionResult(sheets=sheets, markdown=markdown, truncated=truncated)


def _build_column_mapping(
    fields: list[SchemaField], headers: list[str]
) -> dict[str, Optional[str]]:
    """Map schema fields to spreadsheet columns.

    Uses source_column if specified, otherwise tries exact name match
    then case-insensitive match.
    """
    col_map = {}
    header_lower = {h.lower(): h for h in headers}

    for f in fields:
        if f.source_column:
            if f.source_column in headers:
                col_map[f.name] = f.source_column
            elif f.source_column.lower() in header_lower:
                col_map[f.name] = header_lower[f.source_column.lower()]
            else:
                col_map[f.name] = None
        else:
            # Try exact match then case-insensitive
            if f.name in headers:
                col_map[f.name] = f.name
            elif f.name.lower() in header_lower:
                col_map[f.name] = header_lower[f.name.lower()]
            else:
                col_map[f.name] = None

    return col_map


def _coerce_value(
    raw, target_type: str, field_name: str, row_num: int, result: SchemaExtractionResult
):
    """Coerce a cell value to the target schema type."""
    if raw is None:
        return None

    if target_type == "number":
        if isinstance(raw, (int, float)):
            return raw
        try:
            s = str(raw).strip().replace(",", "")
            if "." in s:
                return float(s)
            return int(s)
        except (ValueError, TypeError):
            result.warnings.append(
                "Row {}: cannot convert '{}' to number for field '{}'".format(
                    row_num, raw, field_name
                )
            )
            return None

    if target_type == "boolean":
        if isinstance(raw, bool):
            return raw
        s = str(raw).strip().lower()
        if s in ("true", "yes", "1", "y"):
            return True
        if s in ("false", "no", "0", "n", ""):
            return False
        result.warnings.append(
            "Row {}: cannot convert '{}' to boolean for field '{}'".format(
                row_num, raw, field_name
            )
        )
        return None

    # Default: string
    return str(raw) if raw is not None else None


def _compute_numeric_stats(sheet_data: SheetData) -> dict:
    """Compute min/max/avg stats for numeric columns.

    Returns:
        Dict mapping column name to {min, max, avg, count}.
    """
    stats = {}
    for col_idx, header in enumerate(sheet_data.headers):
        values = []
        for row in sheet_data.rows:
            if col_idx < len(row):
                v = row[col_idx]
                if (
                    isinstance(v, (int, float))
                    and not isinstance(v, bool)
                    and not math.isnan(v)
                ):
                    values.append(v)
                elif isinstance(v, str):
                    try:
                        values.append(float(v.replace(",", "")))
                    except (ValueError, TypeError):
                        pass

        if len(values) >= 3:  # Only report stats if enough numeric values
            stats[header] = {
                "min": min(values),
                "max": max(values),
                "avg": sum(values) / len(values),
                "count": len(values),
            }

    return stats
