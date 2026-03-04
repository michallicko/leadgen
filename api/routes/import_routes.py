"""Import API routes: upload CSV, AI mapping, preview, execute."""

import csv
import io
import json

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook

from ..auth import require_auth, resolve_tenant
from ..models import Tag, CustomFieldDefinition, ImportJob, Owner, db
from ..services.csv_mapper import apply_mapping, call_claude_for_mapping
from ..services.dedup import dedup_preview, execute_import
from ..services.llm_logger import log_llm_usage

imports_bp = Blueprint("imports", __name__)

MAX_CSV_SIZE = 10 * 1024 * 1024  # 10 MB

# Bidirectional mapping: Claude AI target names → frontend target field names.
# Claude returns "contact.email_address" but the frontend expects "email", etc.
CLAUDE_TO_FRONTEND = {
    "contact.first_name": "first_name",
    "contact.last_name": "last_name",
    "contact.email_address": "email",
    "contact.email": "email",
    "contact.phone_number": "phone",
    "contact.phone": "phone",
    "contact.mobile": "mobile",
    "contact.job_title": "job_title",
    "contact.linkedin_url": "linkedin_url",
    "contact.notes": "notes",
    "contact.location_city": "location",
    "contact.location_country": "location",
    "contact.seniority_level": "seniority_level",
    "contact.department": "department",
    "contact.contact_source": "contact_source",
    "contact.language": "language",
    "company.name": "company_name",
    "company.domain": "domain",
    "company.industry": "industry",
    "company.hq_city": "location",
    "company.hq_country": "location",
    "company.company_size": "employee_count",
    "company.business_model": "business_model",
    "company.description": "description",
}

# Explicit reverse mapping: frontend field name → Claude dotted format.
# Built manually (not auto-reversed) to handle one-to-many cases correctly.
FRONTEND_TO_CLAUDE = {
    "first_name": "contact.first_name",
    "last_name": "contact.last_name",
    "email": "contact.email_address",
    "phone": "contact.phone_number",
    "mobile": "contact.mobile",
    "job_title": "contact.job_title",
    "linkedin_url": "contact.linkedin_url",
    "notes": "contact.notes",
    "seniority_level": "contact.seniority_level",
    "department": "contact.department",
    "contact_source": "contact.contact_source",
    "language": "contact.language",
    "company_name": "company.name",
    "domain": "company.domain",
    "industry": "company.industry",
    "employee_count": "company.company_size",
    "business_model": "company.business_model",
    "description": "company.description",
    # "location" is ambiguous (could be city or country) — map to city as best guess
    "location": "contact.location_city",
}

# Targets that indicate the column should be skipped (not mapped)
_SKIP_TARGETS = frozenset({"skip", "ignore", "unmapped", ""})


def _translate_mapping_to_claude(mapping):
    """Translate frontend target field names back to Claude dotted format.

    The frontend sends targets like "email", "domain" but apply_mapping()
    expects "contact.email_address", "company.domain", etc.
    Custom field keys (e.g. "email_secondary") are prefixed to
    "contact.custom.email_secondary" or "company.custom.email_secondary"
    based on the CustomFieldDefinition entity_type.

    Accepts either:
      - Frontend array format: [{source_column, target_field, ...}, ...]
      - Claude dict format: {"mappings": [{csv_header, target, ...}, ...]}
    Always returns Claude dict format.
    """
    from flask import g

    # Build a lookup of known custom field keys → entity_type for the current tenant
    tenant_id = getattr(g, "tenant_id", None) or resolve_tenant()
    custom_field_map = {}  # field_key → entity_type
    if tenant_id:
        custom_defs = CustomFieldDefinition.query.filter_by(
            tenant_id=str(tenant_id),
            is_active=True,
        ).all()
        for cfd in custom_defs:
            custom_field_map[cfd.field_key] = cfd.entity_type

    # Normalize input: frontend sends a list, Claude format is a dict with "mappings"
    if isinstance(mapping, list):
        # Convert frontend ColumnMapping[] → Claude mapping dict
        claude_mappings = []
        for col in mapping:
            entry = {
                "csv_header": col.get("source_column", ""),
                "target": col.get("target_field") or col.get("target"),
                "confidence": col.get("confidence", "low"),
                "sample_values": col.get("sample_values", []),
            }
            # Carry custom field metadata so _auto_create_custom_field_defs can use it
            if col.get("is_custom") and col.get("custom_display_name"):
                entry["suggested_custom_field"] = {
                    "field_label": col["custom_display_name"],
                    "field_type": "text",
                }
            claude_mappings.append(entry)
        translated = {"mappings": claude_mappings}
    elif isinstance(mapping, str):
        import json as _json

        translated = _json.loads(mapping)
    else:
        translated = dict(mapping)

    new_mappings = []
    for m in translated.get("mappings", []):
        target = m.get("target")
        # Clear skip/null targets so apply_mapping ignores them
        if not target or (isinstance(target, str) and target.lower() in _SKIP_TARGETS):
            m = dict(m)
            m["target"] = None
        elif "." in target:
            # Already in dotted format (e.g. "contact.custom.X"), pass through
            pass
        elif target in custom_field_map:
            # Known custom field key — prefix with entity_type.custom.
            m = dict(m)
            m["target"] = f"{custom_field_map[target]}.custom.{target}"
        else:
            m = dict(m)
            translated_target = FRONTEND_TO_CLAUDE.get(target, None)
            if translated_target:
                m["target"] = translated_target
            elif target not in _VALID_FRONTEND_TARGETS:
                # Unknown target that's not a standard field — treat as new custom field
                m["target"] = f"contact.custom.{target}"
            else:
                m["target"] = target
        new_mappings.append(m)
    translated["mappings"] = new_mappings
    return translated


def _auto_create_custom_field_defs(tenant_id, mapping):
    """Scan mapping for custom.* targets and auto-create missing definitions.

    Also checks suggested_custom_field on mappings and creates definitions for those.
    """
    needed = {}  # (entity_type, field_key) → field_label

    for m in mapping.get("mappings", []):
        target = m.get("target")
        if target:
            # e.g. "contact.custom.email_secondary"
            parts = target.split(".", 2)
            if len(parts) == 3 and parts[1] == "custom":
                entity_type, _, field_key = parts
                # Prefer user-edited label from suggested_custom_field, fall back to csv_header
                suggestion = m.get("suggested_custom_field") or {}
                label = suggestion.get("field_label") or m.get(
                    "csv_header", field_key.replace("_", " ").title()
                )
                field_type = suggestion.get("field_type", "text")
                needed[(entity_type, field_key)] = {
                    "label": label,
                    "field_type": field_type,
                }

    for (entity_type, field_key), info in needed.items():
        existing = CustomFieldDefinition.query.filter_by(
            tenant_id=tenant_id,
            entity_type=entity_type,
            field_key=field_key,
        ).first()
        if existing:
            if not existing.is_active:
                existing.is_active = True
            continue
        cfd = CustomFieldDefinition(
            tenant_id=tenant_id,
            entity_type=entity_type,
            field_key=field_key,
            field_label=info["label"],
            field_type=info["field_type"],
        )
        db.session.add(cfd)

    if needed:
        db.session.flush()


def _parse_csv_text(text):
    """Parse CSV text, return (headers, rows) where rows are list of dicts."""
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def _parse_xlsx_bytes(raw):
    """Parse XLSX bytes (first sheet), return (headers, rows) where rows are list of dicts."""
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return [], []
    # Build raw headers, then identify which column indices have non-empty names
    raw_headers = [str(h).strip() if h is not None else "" for h in header_row]
    valid_indices = [i for i, h in enumerate(raw_headers) if h]
    headers = [raw_headers[i] for i in valid_indices]
    rows = []
    for row in rows_iter:
        rows.append(
            {
                raw_headers[i]: (
                    str(row[i]) if i < len(row) and row[i] is not None else ""
                )
                for i in valid_indices
            }
        )
    wb.close()
    return headers, rows


def _rows_to_csv_text(headers, rows):
    """Convert headers + list-of-dicts to CSV text for storage."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


# Secondary target normalization for common variations Claude may return
# that don't appear in CLAUDE_TO_FRONTEND (which only maps dotted names).
_TARGET_NORMALIZE = {
    "title": "job_title",
    "company": "company_name",
    "company_name": "company_name",
    "website": "domain",
    "url": "domain",
    "phone_number": "phone",
    "mobile_phone": "mobile",
    "linkedin": "linkedin_url",
    "email_address": "email",
    "first": "first_name",
    "last": "last_name",
    "fname": "first_name",
    "lname": "last_name",
    "size": "employee_count",
    "employees": "employee_count",
    "role": "job_title",
    "position": "job_title",
    "city": "location",
    "country": "location",
    "desc": "description",
}

# Valid frontend target field values (from MappingStep TARGET_OPTIONS)
_VALID_FRONTEND_TARGETS = {
    "first_name",
    "last_name",
    "email",
    "phone",
    "mobile",
    "job_title",
    "linkedin_url",
    "notes",
    "company_name",
    "domain",
    "industry",
    "employee_count",
    "location",
    "description",
}


def _build_upload_response(
    job_id, filename, total_rows, mapping_result, custom_defs, sample_rows=None
):
    """Transform raw Claude mapping result into the shape the frontend expects.

    The frontend UploadResponse expects:
      { job_id, filename, row_count, columns: ColumnMapping[], warnings, custom_field_defs }

    where ColumnMapping is:
      { source_column, target_field, sample_values, confidence, is_custom, custom_display_name? }

    Args:
        sample_rows: list of dicts (CSV rows) to extract sample values from.
            If None, falls back to whatever Claude returned in the mapping.
    """
    columns = []
    for m in mapping_result.get("mappings", []):
        # Skip entries with empty/whitespace source column names (ghost columns from XLSX)
        csv_header_check = (m.get("csv_header") or "").strip()
        if not csv_header_check:
            continue
        target = m.get("target") or None
        raw_confidence = m.get("confidence", 0)
        if isinstance(raw_confidence, (int, float)):
            if raw_confidence >= 0.75:
                confidence = "high"
            elif raw_confidence >= 0.4:
                confidence = "medium"
            else:
                confidence = "low"
        else:
            confidence = str(raw_confidence) if raw_confidence else "low"

        is_custom = bool(target and "custom" in target)
        suggestion = m.get("suggested_custom_field") or {}
        custom_display_name = suggestion.get("field_label") if is_custom else None

        # Translate Claude AI target names to frontend-expected field names.
        # e.g. "company.domain" → "domain", "contact.email_address" → "email"
        frontend_target = target
        if target and not is_custom:
            frontend_target = CLAUDE_TO_FRONTEND.get(target, target)
            # Secondary normalization for bare names Claude sometimes returns
            if frontend_target not in _VALID_FRONTEND_TARGETS:
                frontend_target = _TARGET_NORMALIZE.get(
                    frontend_target.lower(), frontend_target
                )

        # Populate sample values from actual CSV data when available
        csv_header = m.get("csv_header", "")
        if sample_rows and csv_header:
            samples = [
                str(row.get(csv_header, ""))
                for row in sample_rows[:3]
                if row.get(csv_header)
            ]
        else:
            samples = m.get("sample_values", [])

        columns.append(
            {
                "source_column": csv_header,
                "target_field": frontend_target,
                "sample_values": samples,
                "confidence": confidence,
                "is_custom": is_custom,
                "custom_display_name": custom_display_name,
            }
        )

    # Build a lookup from field_key → source_column using the mapping result.
    # Claude mapping entries with "contact.custom.X" or "company.custom.X" targets
    # carry the original CSV header in csv_header.
    custom_source_map = {}  # field_key → csv_header
    for m in mapping_result.get("mappings", []):
        target = m.get("target") or ""
        parts = target.split(".", 2)
        if len(parts) == 3 and parts[1] == "custom":
            custom_source_map[parts[2]] = m.get("csv_header", "")

    custom_field_defs = []
    for cdef in custom_defs or []:
        d = (
            cdef
            if isinstance(cdef, dict)
            else (cdef.to_dict() if hasattr(cdef, "to_dict") else {})
        )
        # Frontend expects display_name (not field_label) and source_column
        custom_field_defs.append(
            {
                "field_key": d.get("field_key", ""),
                "display_name": d.get(
                    "field_label", d.get("display_name", d.get("field_key", ""))
                ),
                "source_column": custom_source_map.get(
                    d.get("field_key", ""), d.get("field_key", "")
                ),
                "entity_type": d.get("entity_type", "contact"),
            }
        )

    return {
        "job_id": str(job_id),
        "filename": filename,
        "row_count": total_rows,
        "columns": columns,
        "warnings": mapping_result.get("warnings", []),
        "custom_field_defs": custom_field_defs,
    }


@imports_bp.route("/api/imports/upload", methods=["POST"])
@require_auth
def upload_csv():
    """Accept CSV file, parse headers + sample rows, call Claude for mapping.

    Expects multipart form with 'file' field.
    Returns import job with AI-generated column mapping.
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    from flask import g

    user_id = g.current_user.id

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".csv") or filename_lower.endswith(".xlsx")):
        return jsonify({"error": "Only CSV and XLSX files are supported"}), 400

    raw = file.read()
    if len(raw) > MAX_CSV_SIZE:
        return jsonify(
            {"error": f"File too large (max {MAX_CSV_SIZE // (1024 * 1024)} MB)"}
        ), 400

    if filename_lower.endswith(".xlsx"):
        headers, rows = _parse_xlsx_bytes(raw)
        text = _rows_to_csv_text(headers, rows)
    else:
        # Try UTF-8 first, fall back to latin-1
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        headers, rows = _parse_csv_text(text)

    if not headers:
        return jsonify({"error": "Could not parse file headers"}), 400

    sample_rows = []
    for row in rows[:5]:
        sample_rows.append({h: row.get(h, "") for h in headers})

    # Fetch existing custom field definitions for AI context
    custom_defs_rows = CustomFieldDefinition.query.filter_by(
        tenant_id=str(tenant_id),
        is_active=True,
    ).all()
    custom_defs = [d.to_dict() for d in custom_defs_rows]

    # Call Claude for AI column mapping
    usage_info = None
    try:
        mapping_result, usage_info = call_claude_for_mapping(
            headers, sample_rows, custom_defs=custom_defs
        )
    except Exception as e:
        mapping_result = {
            "mappings": [],
            "warnings": [f"AI mapping failed: {str(e)}. Manual mapping required."],
            "combine_columns": [],
        }

    # Compute overall confidence
    confidences = [
        m.get("confidence", 0)
        for m in mapping_result.get("mappings", [])
        if m.get("target")
    ]
    avg_confidence = sum(confidences) / len(confidences) if confidences else 0

    # Create import job
    job = ImportJob(
        tenant_id=str(tenant_id),
        user_id=str(user_id),
        filename=file.filename,
        file_size_bytes=len(raw),
        total_rows=len(rows),
        headers=json.dumps(headers) if isinstance(headers, list) else headers,
        sample_rows=json.dumps(sample_rows)
        if isinstance(sample_rows, list)
        else sample_rows,
        raw_csv=text,
        column_mapping=json.dumps(mapping_result)
        if isinstance(mapping_result, dict)
        else mapping_result,
        mapping_confidence=round(avg_confidence, 2),
        status="mapped",
    )
    db.session.add(job)
    db.session.flush()

    # Log LLM usage if we got usage info from Claude
    if usage_info:
        log_llm_usage(
            tenant_id=str(tenant_id),
            operation="csv_column_mapping",
            model=usage_info["model"],
            input_tokens=usage_info["input_tokens"],
            output_tokens=usage_info["output_tokens"],
            user_id=str(user_id),
            duration_ms=usage_info.get("duration_ms"),
            metadata={
                "import_job_id": str(job.id),
                "filename": file.filename,
                "headers_count": len(headers),
            },
        )

    # Auto-create custom field definitions from mapping
    _auto_create_custom_field_defs(str(tenant_id), mapping_result)

    db.session.commit()

    # Fetch updated custom field defs (including any just created)
    updated_custom_defs = CustomFieldDefinition.query.filter_by(
        tenant_id=str(tenant_id),
        is_active=True,
    ).all()

    return jsonify(
        _build_upload_response(
            job.id,
            file.filename,
            len(rows),
            mapping_result,
            [d.to_dict() for d in updated_custom_defs],
            sample_rows=sample_rows,
        )
    ), 201


@imports_bp.route("/api/imports/<job_id>/remap", methods=["POST"])
@require_auth
def remap_import(job_id):
    """Re-run AI column mapping on an existing import job's CSV data.

    Uses the current prompt (with custom field support) to generate fresh suggestions.
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    job = ImportJob.query.filter_by(id=job_id, tenant_id=str(tenant_id)).first()
    if not job:
        return jsonify({"error": "Import job not found"}), 404

    if job.status == "completed":
        return jsonify({"error": "Cannot remap a completed import"}), 400

    headers, rows = _parse_csv_text(job.raw_csv)
    if not headers:
        return jsonify({"error": "Could not parse stored CSV"}), 400

    sample_rows = [{h: row.get(h, "") for h in headers} for row in rows[:5]]

    custom_defs_rows = CustomFieldDefinition.query.filter_by(
        tenant_id=str(tenant_id),
        is_active=True,
    ).all()
    custom_defs = [d.to_dict() for d in custom_defs_rows]

    try:
        mapping_result, usage_info = call_claude_for_mapping(
            headers, sample_rows, custom_defs=custom_defs
        )
    except Exception as e:
        return jsonify({"error": f"AI mapping failed: {str(e)}"}), 500

    confidences = [
        m.get("confidence", 0)
        for m in mapping_result.get("mappings", [])
        if m.get("target")
    ]
    avg_confidence = sum(confidences) / len(confidences) if confidences else 0

    job.column_mapping = json.dumps(mapping_result)
    job.mapping_confidence = round(avg_confidence, 2)
    job.status = "mapped"
    db.session.flush()

    if usage_info:
        from flask import g

        log_llm_usage(
            tenant_id=str(tenant_id),
            operation="csv_column_remap",
            model=usage_info["model"],
            input_tokens=usage_info["input_tokens"],
            output_tokens=usage_info["output_tokens"],
            user_id=str(g.current_user.id),
            duration_ms=usage_info.get("duration_ms"),
            metadata={
                "import_job_id": str(job.id),
                "filename": job.filename,
                "headers_count": len(headers),
            },
        )

    # Auto-create custom field definitions from mapping
    _auto_create_custom_field_defs(str(tenant_id), mapping_result)

    db.session.commit()

    # Fetch updated custom field defs
    updated_custom_defs = CustomFieldDefinition.query.filter_by(
        tenant_id=str(tenant_id),
        is_active=True,
    ).all()

    # Parse sample rows from stored data for sample_values display
    stored_samples = ImportJob._parse_jsonb(job.sample_rows) or []

    return jsonify(
        _build_upload_response(
            job.id,
            job.filename,
            job.total_rows,
            mapping_result,
            [d.to_dict() for d in updated_custom_defs],
            sample_rows=stored_samples,
        )
    )


@imports_bp.route("/api/imports/<job_id>/preview", methods=["POST"])
@require_auth
def preview_import(job_id):
    """Accept user-adjusted mapping, return parsed rows with dedup results.

    Body: { "mapping": <adjusted mapping object> }
    Returns first 25 rows with dedup status.
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    job = ImportJob.query.filter_by(id=job_id, tenant_id=str(tenant_id)).first()
    if not job:
        return jsonify({"error": "Import job not found"}), 404

    body = request.get_json(silent=True) or {}
    mapping = body.get("mapping")
    if mapping:
        # Translate frontend field names back to Claude dotted format for storage
        mapping = _translate_mapping_to_claude(mapping)
        job.column_mapping = (
            json.dumps(mapping) if isinstance(mapping, dict) else mapping
        )
        db.session.flush()
    else:
        mapping = (
            json.loads(job.column_mapping)
            if isinstance(job.column_mapping, str)
            else job.column_mapping
        )

    # Parse all rows
    headers, all_rows = _parse_csv_text(job.raw_csv)

    # Apply mapping to first 25 rows for preview
    preview_rows = all_rows[:25]
    parsed = [apply_mapping(row, mapping) for row in preview_rows]

    # Run dedup preview
    dedup_results = dedup_preview(str(tenant_id), parsed)

    # Summary counts
    new_contacts = sum(1 for r in dedup_results if r["contact_status"] == "new")
    dup_contacts = sum(1 for r in dedup_results if r["contact_status"] == "duplicate")
    new_companies = sum(1 for r in dedup_results if r["company_status"] == "new")
    existing_companies = sum(
        1 for r in dedup_results if r["company_status"] == "existing"
    )

    job.status = "previewed"
    db.session.commit()

    # Transform dedup results into frontend PreviewRow format:
    # { row_number, data: { first_name, last_name, email, company_name, ... }, status, match_type }
    frontend_rows = []
    for i, r in enumerate(dedup_results):
        contact = r.get("contact", {})
        company = r.get("company", {})
        # Flatten contact + company into a single data dict
        data = {}
        for k, v in contact.items():
            if k.startswith("custom."):
                data[k] = v
            else:
                data[k] = v
        # Add company fields with company_ prefix where needed
        if company.get("name"):
            data["company_name"] = company["name"]
        if company.get("domain"):
            data["company_domain"] = company["domain"]
        for k, v in company.items():
            if k not in ("name", "domain") and v:
                data[f"company_{k}"] = v

        frontend_rows.append(
            {
                "row_number": i + 1,
                "data": data,
                "status": "duplicate"
                if r.get("contact_status") == "duplicate"
                else "new",
                "match_type": r.get("contact_match_type"),
                "match_details": r.get("company_match_type"),
            }
        )

    return jsonify(
        {
            "job_id": str(job.id),
            "preview_rows": frontend_rows,
            "total_rows": job.total_rows,
            "preview_count": len(frontend_rows),
            "new_contacts": new_contacts,
            "duplicates": dup_contacts,
            "updates": 0,
            "new_companies": new_companies,
            "existing_companies": existing_companies,
            "summary": {
                "new_contacts": new_contacts,
                "duplicate_contacts": dup_contacts,
                "new_companies": new_companies,
                "existing_companies": existing_companies,
            },
        }
    )


@imports_bp.route("/api/imports/<job_id>/execute", methods=["POST"])
@require_auth
def execute_import_job(job_id):
    """Create tag, insert companies/contacts, run dedup.

    Body: { "tag_name": "...", "owner_id": "...", "dedup_strategy": "skip"|"update"|"create_new" }
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    job = ImportJob.query.filter_by(id=job_id, tenant_id=str(tenant_id)).first()
    if not job:
        return jsonify({"error": "Import job not found"}), 404

    if job.status == "completed":
        return jsonify({"error": "Import already executed"}), 400

    body = request.get_json(silent=True) or {}
    tag_name = body.get("tag_name", f"import-{job.filename}")
    owner_id = body.get("owner_id")
    strategy = body.get("dedup_strategy", "skip")

    if strategy not in ("skip", "update", "create_new"):
        return jsonify({"error": "Invalid dedup_strategy"}), 400

    # Validate owner if provided
    if owner_id:
        owner = Owner.query.filter_by(id=owner_id, tenant_id=str(tenant_id)).first()
        if not owner:
            return jsonify({"error": "Owner not found"}), 404

    # Create or find tag
    tag = Tag.query.filter_by(tenant_id=str(tenant_id), name=tag_name).first()
    if not tag:
        tag = Tag(tenant_id=str(tenant_id), name=tag_name, is_active=True)
        db.session.add(tag)
        db.session.flush()

    job.tag_id = str(tag.id)
    job.owner_id = str(owner_id) if owner_id else None
    job.dedup_strategy = strategy
    job.status = "importing"
    db.session.flush()

    try:
        # Parse all rows and apply mapping
        mapping = (
            json.loads(job.column_mapping)
            if isinstance(job.column_mapping, str)
            else job.column_mapping
        )

        # Auto-create custom field definitions for any custom.* targets
        _auto_create_custom_field_defs(str(tenant_id), mapping)

        _headers, all_rows = _parse_csv_text(job.raw_csv)
        parsed = [apply_mapping(row, mapping) for row in all_rows]

        # Execute import
        result = execute_import(
            tenant_id=str(tenant_id),
            parsed_rows=parsed,
            tag_id=tag.id,
            owner_id=owner_id,
            import_job_id=job.id,
            strategy=strategy,
        )

        counts = result["counts"]
        dedup_rows = result["dedup_rows"]
        total_conflicts = sum(len(r.get("conflicts", [])) for r in dedup_rows)

        job.contacts_created = counts["contacts_created"]
        job.contacts_updated = counts["contacts_updated"]
        job.contacts_skipped = counts["contacts_skipped"]
        job.companies_created = counts["companies_created"]
        job.companies_linked = counts["companies_linked"]
        job.dedup_results = json.dumps(
            {
                "summary": {
                    "contacts_created": counts["contacts_created"],
                    "contacts_skipped": counts["contacts_skipped"],
                    "contacts_updated": counts["contacts_updated"],
                    "total_conflicts": total_conflicts,
                },
                "rows": dedup_rows,
            }
        )
        job.status = "completed"
        db.session.commit()

        return jsonify(
            {
                "job_id": str(job.id),
                "status": "completed",
                "tag_name": tag_name,
                "counts": counts,
            }
        )

    except Exception as e:
        db.session.rollback()
        job.status = "error"
        job.error = str(e)
        db.session.commit()
        return jsonify({"error": f"Import failed: {str(e)}"}), 500


@imports_bp.route("/api/imports/<job_id>/results", methods=["GET"])
@require_auth
def import_results(job_id):
    """Get detailed dedup results for a completed import.

    Query params:
      filter: all|created|skipped|updated|conflicts (default: all)
      page: 1-based page number (default: 1)
    Returns summary + paginated rows.
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    job = ImportJob.query.filter_by(id=job_id, tenant_id=str(tenant_id)).first()
    if not job:
        return jsonify({"error": "Import job not found"}), 404

    if job.status != "completed":
        return jsonify({"error": "Import not yet completed"}), 400

    raw = job.dedup_results
    if isinstance(raw, str):
        dedup_data = json.loads(raw) if raw else {}
    else:
        dedup_data = raw or {}

    summary = dedup_data.get("summary", {})
    all_rows = dedup_data.get("rows", [])

    filter_type = request.args.get("filter", "all")
    page = max(1, int(request.args.get("page", 1)))
    per_page = 50

    if filter_type == "created":
        filtered = [r for r in all_rows if r.get("action") == "created"]
    elif filter_type == "skipped":
        filtered = [r for r in all_rows if r.get("action") == "skipped"]
    elif filter_type == "updated":
        filtered = [r for r in all_rows if r.get("action") == "updated"]
    elif filter_type == "conflicts":
        filtered = [r for r in all_rows if r.get("conflicts")]
    else:
        filtered = all_rows

    total = len(filtered)
    start = (page - 1) * per_page
    page_rows = filtered[start : start + per_page]

    return jsonify(
        {
            "job_id": str(job.id),
            "summary": summary,
            "filter": filter_type,
            "page": page,
            "per_page": per_page,
            "total": total,
            "rows": page_rows,
        }
    )


@imports_bp.route("/api/imports/<job_id>/status", methods=["GET"])
@require_auth
def import_status(job_id):
    """Get import job status.

    Returns { status, mapping: ColumnMapping[] | null, preview } for the
    frontend ImportStatusResponse type.
    """
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    job = ImportJob.query.filter_by(id=job_id, tenant_id=str(tenant_id)).first()
    if not job:
        return jsonify({"error": "Import job not found"}), 404

    # Build mapping in the ColumnMapping[] format the frontend expects
    mapping = None
    if job.status in ("uploaded", "mapped", "previewed"):
        raw = ImportJob._parse_jsonb(job.column_mapping) or {}
        stored_samples = ImportJob._parse_jsonb(job.sample_rows) or []
        resp = _build_upload_response(
            job.id,
            job.filename,
            job.total_rows,
            raw,
            [],
            sample_rows=stored_samples,
        )
        mapping = resp["columns"]

    return jsonify(
        {
            "status": job.status,
            "mapping": mapping,
            "preview": None,  # preview is re-generated on demand
        }
    )


@imports_bp.route("/api/imports", methods=["GET"])
@require_auth
def list_imports():
    """List past import jobs for tenant."""
    tenant_id = resolve_tenant()
    if not tenant_id:
        return jsonify({"error": "Tenant not found"}), 404

    jobs = (
        ImportJob.query.filter_by(
            tenant_id=str(tenant_id),
        )
        .order_by(ImportJob.created_at.desc())
        .limit(50)
        .all()
    )

    return jsonify(
        {
            "imports": [j.to_dict() for j in jobs],
        }
    )
