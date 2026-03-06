"""Tests for Excel/CSV processor (BL-267)."""

import io
import os
import tempfile

import pytest


class TestCSVProcessing:
    """Tests for CSV file handling."""

    def test_read_csv_from_text_basic(self):
        from api.services.multimodal.excel_processor import read_csv_from_text

        csv_text = "Name,Age,City\nAlice,30,Prague\nBob,25,Berlin\n"
        result = read_csv_from_text(csv_text)

        assert result.headers == ["Name", "Age", "City"]
        assert result.row_count == 2
        assert result.col_count == 3
        assert result.rows[0] == ["Alice", "30", "Prague"]
        assert result.rows[1] == ["Bob", "25", "Berlin"]

    def test_read_csv_from_text_empty(self):
        from api.services.multimodal.excel_processor import read_csv_from_text

        result = read_csv_from_text("")
        assert result.headers == []
        assert result.row_count == 0

    def test_read_csv_from_bytes_utf8(self):
        from api.services.multimodal.excel_processor import read_csv_from_bytes

        data = "Name,Value\nFoo,100\n".encode("utf-8")
        result = read_csv_from_bytes(data)
        assert result.headers == ["Name", "Value"]
        assert result.row_count == 1

    def test_read_csv_from_bytes_latin1(self):
        from api.services.multimodal.excel_processor import read_csv_from_bytes

        # Latin-1 encoded text with special chars
        data = "Name,City\nJan,\xc5\xbdilina\n".encode("latin-1")
        result = read_csv_from_bytes(data)
        assert result.row_count == 1

    def test_csv_file_extraction(self):
        from api.services.multimodal.excel_processor import extract_from_file

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8"
        ) as f:
            f.write("Col1,Col2,Col3\n")
            for i in range(10):
                f.write("a{},b{},c{}\n".format(i, i, i))
            tmp_path = f.name

        try:
            result = extract_from_file(tmp_path)
            assert not result.errors
            assert len(result.sheets) == 1
            assert result.sheets[0].row_count == 10
            assert "Col1" in result.markdown
            assert "a0" in result.markdown
        finally:
            os.unlink(tmp_path)

    def test_csv_auto_header_names(self):
        from api.services.multimodal.excel_processor import read_csv_from_text

        csv_text = ",Value,\n1,2,3\n"
        result = read_csv_from_text(csv_text)
        # Empty headers get auto-named
        assert "col_0" in result.headers
        assert "Value" in result.headers
        assert "col_2" in result.headers


class TestSheetToMarkdown:
    """Tests for markdown table conversion."""

    def test_small_sheet_full_table(self):
        from api.services.multimodal.excel_processor import SheetData, sheet_to_markdown

        data = SheetData(
            name="Test",
            headers=["Name", "Age"],
            rows=[["Alice", 30], ["Bob", 25]],
            row_count=2,
            col_count=2,
        )
        md = sheet_to_markdown(data)
        assert "| Name | Age |" in md
        assert "| --- | --- |" in md
        assert "| Alice | 30 |" in md
        assert "| Bob | 25 |" in md

    def test_empty_headers(self):
        from api.services.multimodal.excel_processor import SheetData, sheet_to_markdown

        data = SheetData(name="Empty", headers=[], rows=[], row_count=0, col_count=0)
        assert sheet_to_markdown(data) == ""

    def test_none_values(self):
        from api.services.multimodal.excel_processor import SheetData, sheet_to_markdown

        data = SheetData(
            name="Test",
            headers=["A", "B"],
            rows=[[None, "val"], ["val", None]],
            row_count=2,
            col_count=2,
        )
        md = sheet_to_markdown(data)
        assert "| val |" in md
        assert "|  |" in md or "| |" in md


class TestSheetToSummary:
    """Tests for large sheet summary generation."""

    def test_summary_includes_stats(self):
        from api.services.multimodal.excel_processor import SheetData, sheet_to_summary

        rows = [[i, i * 10, "text{}".format(i)] for i in range(100)]
        data = SheetData(
            name="BigSheet",
            headers=["ID", "Value", "Label"],
            rows=rows,
            row_count=100,
            col_count=3,
        )
        summary = sheet_to_summary(data)

        assert "BigSheet" in summary
        assert "100 rows" in summary
        assert "3 columns" in summary
        assert "Sample rows" in summary
        assert "Last 3 rows" in summary
        # Should have numeric stats for ID and Value
        assert "min=" in summary
        assert "max=" in summary
        assert "avg=" in summary

    def test_summary_sample_rows(self):
        from api.services.multimodal.excel_processor import SheetData, sheet_to_summary

        rows = [["row_{}".format(i), i] for i in range(60)]
        data = SheetData(
            name="Data",
            headers=["Name", "Num"],
            rows=rows,
            row_count=60,
            col_count=2,
        )
        summary = sheet_to_summary(data)
        # First 5 rows should be present
        assert "row_0" in summary
        assert "row_4" in summary
        # Last 3 rows should be present
        assert "row_57" in summary
        assert "row_59" in summary


class TestSchemaExtraction:
    """Tests for schema-based data extraction."""

    def _create_csv_file(self, content):
        f = tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8"
        )
        f.write(content)
        f.close()
        return f.name

    def test_basic_schema_extraction(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file(
            "Name,Email,Revenue\nAlice,a@b.com,1000\nBob,b@c.com,2000\n"
        )
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[
                    {"name": "name", "type": "string", "source_column": "Name"},
                    {"name": "email", "type": "string", "source_column": "Email"},
                    {"name": "revenue", "type": "number", "source_column": "Revenue"},
                ],
            )
            assert result.error is None
            assert len(result.rows) == 2
            assert result.rows[0]["name"] == "Alice"
            assert result.rows[0]["email"] == "a@b.com"
            assert result.rows[0]["revenue"] == 1000
            assert result.rows[1]["revenue"] == 2000
        finally:
            os.unlink(path)

    def test_case_insensitive_mapping(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("NAME,email\nAlice,a@b.com\n")
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[
                    {"name": "name", "source_column": "name"},
                    {"name": "email", "source_column": "EMAIL"},
                ],
            )
            assert result.error is None
            assert result.rows[0]["name"] == "Alice"
            assert result.rows[0]["email"] == "a@b.com"
        finally:
            os.unlink(path)

    def test_unmapped_columns(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("A,B,C,D\n1,2,3,4\n")
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[{"name": "A"}, {"name": "B"}],
            )
            assert result.error is None
            assert "C" in result.unmapped_columns
            assert "D" in result.unmapped_columns
        finally:
            os.unlink(path)

    def test_type_coercion_number(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("Val\n42\n3.14\nnot_a_number\n")
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[{"name": "Val", "type": "number"}],
            )
            assert result.rows[0]["Val"] == 42
            assert result.rows[1]["Val"] == 3.14
            assert result.rows[2]["Val"] is None
            assert any("cannot convert" in w for w in result.warnings)
        finally:
            os.unlink(path)

    def test_type_coercion_boolean(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("Flag\ntrue\nfalse\nyes\nno\nmaybe\n")
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[{"name": "Flag", "type": "boolean"}],
            )
            assert result.rows[0]["Flag"] is True
            assert result.rows[1]["Flag"] is False
            assert result.rows[2]["Flag"] is True
            assert result.rows[3]["Flag"] is False
            assert result.rows[4]["Flag"] is None  # "maybe" can't be coerced
        finally:
            os.unlink(path)

    def test_missing_source_column(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("A,B\n1,2\n")
        try:
            result = extract_data_with_schema(
                path,
                schema_fields=[{"name": "C", "source_column": "NonExistent"}],
            )
            assert result.rows[0]["C"] is None
            assert any("no source column" in w for w in result.warnings)
        finally:
            os.unlink(path)

    def test_empty_file(self):
        from api.services.multimodal.excel_processor import extract_data_with_schema

        path = self._create_csv_file("")
        try:
            result = extract_data_with_schema(path, schema_fields=[{"name": "A"}])
            assert result.error is not None
            assert "No headers" in result.error
        finally:
            os.unlink(path)


class TestExcelProcessing:
    """Tests for Excel (.xlsx) file handling."""

    def test_extract_real_xlsx(self):
        """Test with a real .xlsx created in-memory."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from api.services.multimodal.excel_processor import extract_from_bytes

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product", "Price", "Quantity"])
        ws.append(["Widget", 9.99, 100])
        ws.append(["Gadget", 19.99, 50])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = extract_from_bytes(xlsx_bytes, filename="products.xlsx")
        assert not result.errors
        assert len(result.sheets) >= 1
        assert "Product" in result.markdown
        assert "Widget" in result.markdown
        assert "9.99" in result.markdown

    def test_multi_sheet_discovery(self):
        """Test sheet discovery on multi-sheet workbook."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from api.services.multimodal.excel_processor import extract_from_bytes

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Month", "Revenue"])
        ws1.append(["Jan", 1000])

        ws2 = wb.create_sheet("Costs")
        ws2.append(["Category", "Amount"])
        ws2.append(["Rent", 500])

        buf = io.BytesIO()
        wb.save(buf)

        result = extract_from_bytes(buf.getvalue(), filename="report.xlsx")
        assert not result.errors
        sheet_names = [s.name for s in result.sheets]
        assert "Sales" in sheet_names
        assert "Costs" in sheet_names
        assert "Workbook Sheets" in result.markdown

    def test_large_sheet_gets_summary(self):
        """Sheets with >=50 rows should get summary, not full table."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from api.services.multimodal.excel_processor import extract_from_bytes

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BigData"
        ws.append(["ID", "Value", "Category"])
        for i in range(80):
            ws.append([i, i * 10, "cat_{}".format(i % 5)])

        buf = io.BytesIO()
        wb.save(buf)

        result = extract_from_bytes(buf.getvalue(), filename="big.xlsx")
        assert not result.errors
        assert "80 rows" in result.markdown
        assert "3 columns" in result.markdown
        assert "Sample rows" in result.markdown
        assert "min=" in result.markdown

    def test_schema_extraction_xlsx(self):
        """Test schema-based extraction on xlsx."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        from api.services.multimodal.excel_processor import extract_data_with_schema

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Full Name", "Email", "Revenue"])
        ws.append(["Alice Smith", "alice@co.com", 5000])
        ws.append(["Bob Jones", "bob@co.com", 3000])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            tmp_path = f.name

        try:
            result = extract_data_with_schema(
                tmp_path,
                schema_fields=[
                    {"name": "name", "type": "string", "source_column": "Full Name"},
                    {"name": "revenue", "type": "number", "source_column": "Revenue"},
                ],
            )
            assert result.error is None
            assert len(result.rows) == 2
            assert result.rows[0]["name"] == "Alice Smith"
            assert result.rows[0]["revenue"] == 5000
            assert "Email" in result.unmapped_columns
        finally:
            os.unlink(tmp_path)


class TestTokenBudget:
    """Tests for output truncation."""

    def test_truncation_on_large_content(self):
        from api.services.multimodal.excel_processor import (
            read_csv_from_text,
            MAX_OUTPUT_CHARS,
        )

        # Create a CSV with <50 rows but very long values to exceed the budget
        lines = ["Col1,Col2,Col3,Col4,Col5"]
        for i in range(40):
            lines.append(
                "{0}_aaaa_{0}_aaaa_{0}_aaaa_{0}_aaaa_{0}_aaaa,"
                "{0}_bbbb_{0}_bbbb_{0}_bbbb_{0}_bbbb_{0}_bbbb,"
                "{0}_cccc_{0}_cccc_{0}_cccc_{0}_cccc_{0}_cccc,"
                "{0}_dddd_{0}_dddd_{0}_dddd_{0}_dddd_{0}_dddd,"
                "{0}_eeee_{0}_eeee_{0}_eeee_{0}_eeee_{0}_eeee".format("x" * 30)
            )
        csv_text = "\n".join(lines)

        from api.services.multimodal.excel_processor import _build_result_from_sheet

        data = read_csv_from_text(csv_text)
        result = _build_result_from_sheet(data)

        # Should be truncated since 40 rows * very long values > 8000 chars
        assert result.truncated
        assert "[Content truncated]" in result.markdown
        assert (
            len(result.markdown) <= MAX_OUTPUT_CHARS + 50
        )  # small buffer for truncation notice


class TestNumericStats:
    """Tests for numeric column statistics."""

    def test_stats_computation(self):
        from api.services.multimodal.excel_processor import (
            SheetData,
            _compute_numeric_stats,
        )

        data = SheetData(
            name="Test",
            headers=["Name", "Score", "Grade"],
            rows=[
                ["Alice", 85, "A"],
                ["Bob", 92, "A"],
                ["Carol", 78, "B"],
                ["Dave", 65, "C"],
                ["Eve", 90, "A"],
            ],
            row_count=5,
            col_count=3,
        )
        stats = _compute_numeric_stats(data)
        assert "Score" in stats
        assert stats["Score"]["min"] == 65
        assert stats["Score"]["max"] == 92
        assert stats["Score"]["count"] == 5
        # Grade should not appear (not numeric)
        assert "Grade" not in stats

    def test_stats_with_mixed_values(self):
        from api.services.multimodal.excel_processor import (
            SheetData,
            _compute_numeric_stats,
        )

        data = SheetData(
            name="Test",
            headers=["Val"],
            rows=[[10], ["20"], [None], ["abc"], [30]],
            row_count=5,
            col_count=1,
        )
        stats = _compute_numeric_stats(data)
        assert "Val" in stats
        assert stats["Val"]["min"] == 10
        assert stats["Val"]["max"] == 30
        assert stats["Val"]["count"] == 3
